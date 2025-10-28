import openpyxl
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.protection import WorkbookProtection
from services.report_chart_interactive import export_interactive_chart
from datetime import datetime

def export_monthly_report_to_excel(report, output_path, month_name, year):
    wb = openpyxl.Workbook()

    # === SHEET 1: SUMMARY ===
    summary_ws = wb.active
    summary_ws.title = "Summary"

    summary_ws.merge_cells("A1:C1")
    summary_ws["A1"] = f"Monthly Report: {month_name} {year}"
    summary_ws["A1"].font = Font(size=14, bold=True)
    summary_ws["A1"].alignment = Alignment(horizontal="center")

    # Timestamp
    summary_ws["A2"] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    summary_ws["A2"].font = Font(italic=True, size=10, color="555555")
    summary_ws["A2"].alignment = Alignment(horizontal="center")
    summary_ws.merge_cells("A2:C2")

    summary = [
        ("Total Spent", report.total, "currency"),
        ("Total Income", report.total_monthly_income, "currency"),
        ("Income Used (%)", report.income_usage_percent / 100, "percent"),
        ("Income Saved", report.income_saved, "currency"),
        ("Budget Cap", report.budget_cap, "currency"),
        ("Budget Used (%)", report.budget_used_percent / 100, "percent"),
        ("Days Covered", report.days_covered, None),
        ("Avg/Day", report.avg_per_day, "currency"),
        ("Projected Total", report.projected_total, "currency"),
        ("Remaining Budget", report.remaining_budget, "currency"),
        ("On Track", "Yes" if report.is_on_track else "No", "ontrack"),
        ("Transaction Count", report.transaction_count, None),
    ]

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    start_row = 3
    for i, (label, value, fmt) in enumerate(summary, start=start_row):
        summary_ws[f"A{i}"] = label
        summary_ws[f"B{i}"] = value
        summary_ws[f"A{i}"].font = Font(bold=True)
        if fmt == "currency":
            summary_ws[f"B{i}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        elif fmt == "percent":
            summary_ws[f"B{i}"].number_format = "0.00%"
        elif fmt == "ontrack":
            fill = PatternFill(
                start_color="C8E6C9" if value == "Yes" else "FFCDD2",
                end_color="C8E6C9" if value == "Yes" else "FFCDD2",
                fill_type="solid"
            )
            summary_ws[f"B{i}"].fill = fill
        summary_ws[f"A{i}"].border = thin_border
        summary_ws[f"B{i}"].border = thin_border

    # -------- Income breakdown --------
    row_offset = start_row + len(summary) + 2
    header_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")

    summary_ws.append([])  # spacer
    summary_ws[f"A{row_offset}"] = "Income Source"
    summary_ws[f"B{row_offset}"] = "Amount ($)"
    summary_ws[f"C{row_offset}"] = "% of Income"
    for col in "ABC":
        summary_ws[f"{col}{row_offset}"].font = Font(bold=True)
        summary_ws[f"{col}{row_offset}"].fill = header_fill
        summary_ws[f"{col}{row_offset}"].border = thin_border

    # write income rows
    income_start = row_offset + 1
    cur = income_start
    for src, amt in report.income_sources.items():
        summary_ws[f"A{cur}"] = src
        summary_ws[f"B{cur}"] = amt
        summary_ws[f"B{cur}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        summary_ws[f"C{cur}"] = report.income_source_percentages.get(src, 0) / 100
        summary_ws[f"C{cur}"].number_format = "0.00%"
        for col in "ABC":
            summary_ws[f"{col}{cur}"].border = thin_border
        cur += 1

    income_count = cur - income_start
    if income_count > 0:
        income_end = cur - 1
        income_total_row = income_end + 1
        summary_ws[f"A{income_total_row}"] = "Grand Total"
        summary_ws[f"A{income_total_row}"].font = Font(bold=True)
        summary_ws[f"A{income_total_row}"].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        # safe SUM that never includes itself
        summary_ws[f"B{income_total_row}"] = f"=SUM(B{income_start}:B{income_end})"
        summary_ws[f"B{income_total_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        summary_ws[f"B{income_total_row}"].font = Font(bold=True)
        summary_ws[f"B{income_total_row}"].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        summary_ws[f"C{income_total_row}"] = ""
        for col in "ABC":
            summary_ws[f"{col}{income_total_row}"].border = thin_border
    else:
        # no income rows: write a single “No income” line (optional)
        income_total_row = income_start
        summary_ws[f"A{income_total_row}"] = "No income sources"
        for col in "ABC":
            summary_ws[f"{col}{income_total_row}"].border = thin_border

    # -------- Category breakdown --------
    cat_header_row = income_total_row + 2
    summary_ws[f"A{cat_header_row}"] = "Category"
    summary_ws[f"B{cat_header_row}"] = "Total ($)"
    summary_ws[f"C{cat_header_row}"] = "% of Spending"
    for col in "ABC":
        summary_ws[f"{col}{cat_header_row}"].font = Font(bold=True)
        summary_ws[f"{col}{cat_header_row}"].fill = header_fill
        summary_ws[f"{col}{cat_header_row}"].border = thin_border

    cat_start = cat_header_row + 1
    cur = cat_start
    for cat, amt in report.category_totals.items():
        summary_ws[f"A{cur}"] = cat.capitalize()
        summary_ws[f"B{cur}"] = amt
        summary_ws[f"B{cur}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        summary_ws[f"C{cur}"] = report.category_percentages.get(cat, 0) / 100
        summary_ws[f"C{cur}"].number_format = "0.00%"
        for col in "ABC":
            summary_ws[f"{col}{cur}"].border = thin_border
        cur += 1

    cat_count = cur - cat_start
    if cat_count > 0:
        cat_end = cur - 1
        cat_total_row = cat_end + 1
        summary_ws[f"A{cat_total_row}"] = "Grand Total"
        summary_ws[f"A{cat_total_row}"].font = Font(bold=True)
        summary_ws[f"A{cat_total_row}"].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        # safe SUM that never includes itself
        summary_ws[f"B{cat_total_row}"] = f"=SUM(B{cat_start}:B{cat_end})"
        summary_ws[f"B{cat_total_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        summary_ws[f"B{cat_total_row}"].font = Font(bold=True)
        summary_ws[f"B{cat_total_row}"].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        summary_ws[f"C{cat_total_row}"] = ""
        for col in "ABC":
            summary_ws[f"{col}{cat_total_row}"].border = thin_border
    else:
        cat_total_row = cat_start
        summary_ws[f"A{cat_total_row}"] = "No categories"
        for col in "ABC":
            summary_ws[f"{col}{cat_total_row}"].border = thin_border

    # Auto width
    for col in summary_ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        summary_ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Protect summary sheet
    summary_ws.protection.sheet = True
    summary_ws.protection.password = "readonly"
    summary_ws.protection.enable()
    wb.security = WorkbookProtection(workbookPassword="readonly", lockStructure=True)

    # === SHEET 2: TRANSACTIONS WITH CATEGORY SUBTOTALS ===
    tx_ws = wb.create_sheet("Transactions")
    tx_ws.append(["Category", "Date", "Description", "Type", "Amount"])
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")

    for cell in tx_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    current_row = 2
    sorted_txs = sorted(report.transactions, key=lambda t: (t.get("category", ""), t["date"]))

    current_cat = None
    category_total = 0.0

    for tx in sorted_txs:
        cat = tx.get("category", "").capitalize()
        if cat != current_cat and current_cat is not None:
            # Write subtotal before switching
            tx_ws.append(["", "", "", f"{current_cat} Total", category_total])
            for col in range(1, 6):
                cell = tx_ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                cell.border = thin_border
            tx_ws[f"E{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            current_row += 2
            category_total = 0.0

        if cat != current_cat:
            tx_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            tx_ws[f"A{current_row}"] = cat
            tx_ws[f"A{current_row}"].font = Font(bold=True, size=12)
            tx_ws[f"A{current_row}"].fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
            current_cat = cat
            current_row += 1

        tx_ws.append([
            cat,
            tx["date"],
            tx["description"],
            tx["type"],
            tx["amount"]
        ])
        category_total += tx["amount"]
        current_row += 1

    # Final category subtotal
    tx_ws.append(["", "", "", f"{current_cat} Total", category_total])
    for col in range(1, 6):
        cell = tx_ws.cell(row=current_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        cell.border = thin_border
    tx_ws[f"E{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Format amount column
    for row in tx_ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    tx_ws.freeze_panes = "A2"

    # Auto width
    for col in tx_ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        tx_ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb.save(output_path)
    export_interactive_chart(report, output_path, month_name, year)
